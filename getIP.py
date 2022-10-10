from pprint import pprint
from openpyxl import load_workbook
import re

xlsFile = "C:/Work/01_Projects/IMS PP/host allocation check/cm_sz_nolinks.xlsx"
wb = load_workbook(xlsFile, data_only=True)


def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return (
        sheet.cell(rng[0].min_row, rng[0].min_col).value
        if len(rng) != 0
        else cell.value
    )


col_list = ["G", "O"]
ip_list = []
ip_record = []

for sheet_ in wb.worksheets:
    for col_ in col_list:
        for cell_ in sheet_[col_]:
            value_ = str(getMergedCellVal(sheet_, cell_))
            try:
                ipaddress.ip_address(value_)
                if value_ not in ip_list:
                    ip_ = str(value_).strip()
                    sys_ = str(
                        getMergedCellVal(
                            sheet_, sheet_.cell(row=cell_.row, column=cell_.column - 5)
                        )
                    )
                    host_ = str(
                        getMergedCellVal(
                            sheet_, sheet_.cell(row=cell_.row, column=cell_.column - 1)
                        )
                    )
                    net_ = re.search(r"\d+\.\d+\.\d+", ip_).group(0)
                    hostip_ = re.search(r"\d+$", ip_).group(0)
                    ip_list.append(ip_)
                    ip_record.append(
                        {
                            "HostSys": sys_,
                            "host": host_,
                            "ip": ip_,
                            "net": net_,
                            "hostip": hostip_,
                        }
                    )
            except ValueError:
                pass

for record in ip_record:
    print(
        record["HostSys"]
        + ","
        + record["host"]
        + ","
        + record["ip"]
        + ","
        + record["net"]
        + ","
        + record["hostip"]
    )
